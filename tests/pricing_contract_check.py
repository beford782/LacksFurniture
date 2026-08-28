#!/usr/bin/env python3
"""Phase 2.1 pricing contract check — shipped-state lock + real-pipeline traversal.

Owns the DARK framework's production/test separation:

  * the committed production pricing configuration is EXACTLY the dark
    contract — enabled false, displayEnabled false, every surface false,
    products [] and formulas [], every business policy explicitly UNAPPROVED
    and unset (no default freshness cadence, no approved source host, an
    unapproved presentation with business / legal / native-language review
    recorded separately) — at every layer (canonical source, workbook Pricing
    tab envelope, generated store-config, generated demo bundle) with ZERO
    transforms between canonical source and shipped config;
  * no numeric price, payment or transaction figure ships in any production
    artifact. DOM silence is a CONTAINMENT proof as of Phase 2.1b (a reviewed
    change to the 2.1a token-absence lock, named in the PR): index.html
    carries exactly one pricing artifact — the pure resolveDarkPricing
    DEFINITION between its two markers, with ZERO live call sites — every
    pricing token in the file lives inside that block, and nothing anywhere
    reads STORE_CONFIG.pricing, so no runtime code path can depend on the
    contract and rendered output is structurally invariant to it
    (tests/pricing_resolver_check.mjs owns the resolver's behaviour);
  * the Daybreak Promotions envelope is untouched (still exactly
    {promotions, financing}) — pricing rides its OWN tab;
  * the price-specific canonical allowlist exists and is narrow, and the
    shipped source policy is unapproved and empty (candidates ≠ approval);
  * a POPULATED, NON-SHIPPING fixture (tests/fixtures/pricing_populated_fixture
    .json) traverses the real pipeline in a temp workbook — canonical envelope
    -> Pricing tab (+ its explicitly authorized financing in the Promotions
    envelope) -> converter build_pricing -> build_store_config ->
    validate_financing + validate_pricing under an INJECTED clock — and is
    admitted; then single-field corruptions covering every reviewed contract
    rule are each REFUSED, so the admission is not vacuous. (The JS resolver
    leg lives in tests/pricing_resolver_check.mjs, which derives its governed
    variants from this same fixture under the same injected clock.)
  * the fixture's placeholder text never appears in a shipping artifact;
  * every mutation-sweep find string this slice adds matches its target
    EXACTLY ONCE (the sweep itself does not check uniqueness).

This lock is a PRE-ACTIVATION lock, permanent until Phase 2.2: relaxing the
displayEnabled / surfaces assertions is the 2.2 activation change and arrives
only with Blake's plus written business/legal approval recorded on the item.
Populating products/formulas or approving a policy is NOT a lock relaxation —
it is the ordinary governed path (validate_pricing admits or refuses it) — but
the assertions below that pin the empty/unapproved state must then be
re-pinned to the approved content in the same reviewed diff. Empty production
data does not close Phase 2.1.

Run: python tests/pricing_contract_check.py
"""

import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO, "tools"))
import validation  # noqa: E402
import workbook_schema as schema  # noqa: E402
import convert_store_data as converter  # noqa: E402

import openpyxl  # noqa: E402

passed = failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1
        print(f"  [ok]   {name}")
    else:
        failed += 1
        print(f"  [FAIL] {name}" + (f" - {detail}" if detail else ""))


def _read(path):
    with io.open(os.path.join(REPO, path), encoding="utf-8", newline="") as f:
        return f.read()


def _load(path):
    return json.loads(_read(path))


def _lf(text):
    return text.replace("\r\n", "\n")


SIZES = ["twin", "twin_xl", "full", "queen", "king", "cal_king"]
UNAPPROVED = {"status": "unapproved", "by": "", "at": None}
DARK = {
    "schemaVersion": 1,
    "enabled": False,
    "displayEnabled": False,
    "currency": "USD",
    "authority": {"owner": "", "role": ""},
    "freshness": {"status": "unapproved", "maxAgeDays": None, "approvedBy": "", "approvedAt": None},
    "sourcePolicy": {"status": "unapproved", "allowedSourceHosts": [], "approvedBy": "", "approvedAt": None},
    "purchaseAssessment": {"policy": "runtime-transaction-amount"},
    "sizes": SIZES,
    "surfaces": {"drawer": False, "sleepSystem": False, "results": False,
                 "handoff": False, "sleepPlan": False},
    "presentation": {
        "status": "unapproved",
        "approvals": {"business": dict(UNAPPROVED), "legal": dict(UNAPPROVED),
                      "nativeReview": {"status": "pending", "by": "", "at": None}},
        "assumptions": [], "disclosures": [], "states": {},
    },
    "products": [],
    "formulas": [],
}
# Strings that would betray a shipped price, payment or transaction figure.
# Deliberately the FIELD names the contract itself uses, so populated data
# cannot slip into production by accident.
NO_SHIP_STRINGS = ('"amountMinor"', '"transactionAmountMinor"', '"transactionAmount"',
                   '"monthlyPayment"', '"perMonth"', '"estimatedPayment"',
                   '"publishedPaymentFactor"', "FIXTURE")
# DOM silence, Phase 2.1b form. ABSENT tokens may appear NOWHERE in the app:
# nothing reads the shipped contract (no STORE_CONFIG.pricing read exists) and
# no accessor or legacy resolver name was introduced. CONTAINED tokens exist
# ONLY inside the marked resolver-definition block — the one pricing artifact
# index.html carries, a pure function with zero live call sites.
DOM_ABSENT = ("STORE_CONFIG.pricing", "getPricingConfig", "resolvePrice",
              "pricing.presentation", "purchaseAssessment")
DOM_CONTAINED = ("resolveDarkPricing", "displayEnabled", "amountMinor",
                 "transactionAmountMinor", "minimumPurchase", "sourcePolicy")
RESOLVER_START = "// ═══ PHASE 2.1B DARK RESOLVER (definition only — zero live call sites) ═══"
RESOLVER_END = "// ═══ END PHASE 2.1B DARK RESOLVER ═══"

# ---- production artifacts ---------------------------------------------------
print("Production shipped-state lock (dark until Phase 2.2):")
src = _load("incoming/lacks_pricing.json")
check("canonical pricing source is exactly the dark contract",
      src.get("pricing") == DARK,
      json.dumps(src.get("pricing"))[:200])
check("canonical pricing source carries _meta documentation (never shipped)",
      isinstance(src.get("_meta"), dict) and set(src) == {"_meta", "pricing"})

cfg = _load("data/store-config.json")
check("generated store-config pricing is exactly the dark contract",
      cfg.get("pricing") == DARK)
check("generated store-config pricing deep-equals the canonical source (zero transforms)",
      cfg.get("pricing") == src.get("pricing"))
check("pricing sits after financing in the shipped key order",
      list(cfg).index("pricing") == list(cfg).index("financing") + 1)
raw_cfg = _read("data/store-config.json")
for s in NO_SHIP_STRINGS:
    check(f"no shipped price/payment/transaction figure: {s!r} absent from store-config", s not in raw_cfg)
pr = cfg.get("pricing", {})
check("shipped pricing products and formulas are empty",
      pr.get("products") == [] and pr.get("formulas") == [])
check("shipped pricing is disabled AND display-disabled AND every surface off",
      pr.get("enabled") is False and pr.get("displayEnabled") is False
      and all(v is False for v in pr.get("surfaces", {}).values())
      and set(pr.get("surfaces", {})) == validation.PRICING_SURFACES)
check("shipped freshness policy is unapproved with NO default maxAgeDays",
      pr.get("freshness", {}).get("status") == "unapproved"
      and pr.get("freshness", {}).get("maxAgeDays") is None)
check("shipped source policy is unapproved with NO approved hosts",
      pr.get("sourcePolicy", {}).get("status") == "unapproved"
      and pr.get("sourcePolicy", {}).get("allowedSourceHosts") == [])
check("shipped presentation is unapproved with separate business/legal/native records",
      pr.get("presentation", {}).get("status") == "unapproved"
      and set(pr.get("presentation", {}).get("approvals", {})) == validation.PRICING_PRESENTATION_APPROVAL_KEYS
      and pr.get("presentation", {}).get("approvals", {}).get("nativeReview", {}).get("status") == "pending")
check("shipped contract carries no combined ES/legal flag and no catalog-wide clearance",
      "esReviewStatus" not in pr and "mapClearance" not in pr)
check("shipped purchase-threshold policy is runtime-transaction-amount only",
      pr.get("purchaseAssessment") == {"policy": "runtime-transaction-amount"})

# DOM silence (2.1b containment form): nothing reads the shipped contract, and
# every pricing token lives inside the marked resolver definition. The demo
# bundle is a copy of index.html and gets the identical scan.
for page in ("index.html", "demo/black-friday/index.html"):
    root = _read(page)
    s_count, e_count = root.count(RESOLVER_START), root.count(RESOLVER_END)
    check(f"{page}: resolver markers appear exactly once each", s_count == 1 and e_count == 1,
          f"start={s_count} end={e_count}")
    if s_count == 1 and e_count == 1:
        s_idx, e_idx = root.index(RESOLVER_START), root.index(RESOLVER_END)
        check(f"{page}: resolver block is well-formed (start before end)", s_idx < e_idx)
        block = root[s_idx:e_idx + len(RESOLVER_END)]
        for tok in DOM_ABSENT:
            check(f"{page}: token absent everywhere (nothing reads the contract): {tok!r}",
                  tok not in root)
        for tok in DOM_CONTAINED:
            n_all, n_block = root.count(tok), block.count(tok)
            check(f"{page}: token {tok!r} appears only inside the resolver block",
                  n_block >= 1 and n_all == n_block, f"file={n_all} block={n_block}")
        # Zero live call sites: the resolver NAME appears exactly once as its
        # own declaration (plus marker-comment mentions inside the block).
        check(f"{page}: resolveDarkPricing is declared once and never called",
              root.count("function resolveDarkPricing(") == 1
              and root.count("resolveDarkPricing(") == 1)
root = _read("index.html")

# The workbook: its own tab, chunked envelope, Daybreak envelope untouched.
wb = openpyxl.load_workbook(os.path.join(REPO, "incoming", "Lacks_Store_Data.xlsx"))
check("workbook carries the Pricing tab as the last schema tab",
      wb.sheetnames == schema.get_tab_names() and wb.sheetnames[-1] == "Pricing")
ws = wb["Pricing"]
payload = "".join(str(row[0].value or "") for row in ws.iter_rows(min_row=2))
envelope = json.loads(payload)
check("workbook Pricing envelope is exactly {pricing}", set(envelope) == {"pricing"})
check("workbook Pricing slot deep-equals the canonical source", envelope.get("pricing") == DARK)
check("_meta is not propagated into the Pricing envelope", "_meta" not in payload)
for s in NO_SHIP_STRINGS:
    check(f"no shipped price/payment/transaction figure: {s!r} absent from the workbook envelope", s not in payload)
pws = wb["Promotions"]
ppayload = "".join(str(row[0].value or "") for row in pws.iter_rows(min_row=2))
prod_envelope = json.loads(ppayload)
check("Daybreak Promotions envelope still carries exactly promotions and financing",
      set(prod_envelope) == {"promotions", "financing"})
check("pricing never entered the Promotions envelope", '"pricing"' not in ppayload)
check("production financing keeps exact-term OUTPUT authorization OFF (the runtime render gate stands)",
      cfg.get("financing", {}).get("exactPromotionsEnabled") is False)

# The demo bundle derives from production config and must carry the same dark block.
demo_cfg = _load("demo/black-friday/data/store-config.json")
check("demo bundle pricing deep-equals production pricing", demo_cfg.get("pricing") == DARK)

# The canonical allowlists: price-specific and narrow; financing list distinct.
hosts = validation.load_source_hosts()
price_hosts = hosts.get("priceSourceHosts")
fin_hosts = hosts.get("financingSourceHosts")
check("tools/source_hosts.json declares priceSourceHosts (a technical ceiling, not approval)",
      isinstance(price_hosts, list) and price_hosts == ["lacks.com", "www.lacks.com"]
      and "not business approval" in hosts.get("_comment", ""))
check("price allowlist admits no lender, archive or image-CDN host",
      not any(h for h in price_hosts if "synchrony" in h or "archive" in h or "linqcdn" in h))
check("financing allowlist is a distinct policy from the price allowlist",
      isinstance(fin_hosts, list) and set(fin_hosts) != set(price_hosts))

# The shipped contract validates clean under the real clock and an injected one.
mattress_ids = {m["id"] for tier in _load("data/mattresses.json").values() for m in tier}
accessory_ids = {a["id"] for a in _load("data/accessories.json")}
KW = dict(allowed_source_hosts=price_hosts, financing_source_hosts=fin_hosts,
          mattress_ids=mattress_ids, accessory_ids=accessory_ids)
rep = validation.validate_pricing(cfg, **KW)
check("shipped pricing validates clean under the real clock (no errors, no warnings)",
      rep.ok and not rep.warnings, "; ".join(rep.errors[:2] + rep.warnings[:2]))

# ---- the operating-state lock names the switch --------------------------------
ci = _lf(_read(".github/workflows/ci.yml"))
check("CI operating-state lock names pricing.displayEnabled",
      "pricing ships dark" in ci and "displayEnabled" in ci)
check("CI runs this suite, the pricing totality suite and the 2.1b resolver suite",
      "tests/pricing_contract_check.py" in ci and "tests/pricing_totality_check.py" in ci
      and "tests/pricing_resolver_check.mjs" in ci)
check("CI protects the canonical pricing source",
      "incoming/lacks_pricing.json" in ci)

# ---- the populated NON-SHIPPING fixture traverses the real pipeline ------------
print("Populated fixture traversal (temp workbook, injected clock):")
fx = _load("tests/fixtures/pricing_populated_fixture.json")
CLOCK = datetime.fromisoformat(fx["_meta"]["clock"])
fx_pricing = fx["pricing"]
fx_fin = fx["financing"]
check("fixture is populated (one product, one formula) and dark for display",
      len(fx_pricing["products"]) == 1 and len(fx_pricing["formulas"]) == 1
      and fx_pricing["displayEnabled"] is False and fx_pricing["enabled"] is True)
check("fixture financing is enabled and current, with exact-term OUTPUT authorization OFF "
      "exactly as production ships it (gate split 2026-08-28: the flag never gates validation)",
      fx_fin.get("exactPromotionsEnabled") is False and fx_fin.get("enabled") is True
      and all(datetime.fromisoformat(p["verifiedAt"]) <= CLOCK for p in fx_fin["plans"] if "verifiedAt" in p))
for rel in ("data/store-config.json", "incoming/lacks_pricing.json",
            "demo/black-friday/data/store-config.json", "incoming/lacks_financing.json"):
    check(f"fixture placeholder text never appears in {rel}", "FIXTURE" not in _read(rel))
check("fixture placeholder text never appears in the workbook envelopes",
      "FIXTURE" not in payload and "FIXTURE" not in ppayload)


def _fixture_workbook(dst, pricing_value, financing_value=None, chunk=97):
    """Copy the committed workbook and rewrite its Pricing tab (and, when
    given, the financing half of the Promotions envelope) with the fixture,
    chunked small so reassembly is genuinely exercised."""
    shutil.copyfile(os.path.join(REPO, "incoming", "Lacks_Store_Data.xlsx"), dst)
    w = openpyxl.load_workbook(dst)
    s = w["Pricing"]
    s.delete_rows(2, s.max_row)
    text = json.dumps({"pricing": pricing_value}, ensure_ascii=False, separators=(",", ":"))
    rows = 0
    for i in range(0, len(text), chunk):
        s.append([text[i:i + chunk]])
        rows += 1
    if financing_value is not None:
        ps = w["Promotions"]
        ps.delete_rows(2, ps.max_row)
        env = json.dumps({"promotions": prod_envelope["promotions"], "financing": financing_value},
                         ensure_ascii=False, separators=(",", ":"))
        for i in range(0, len(env), 30000):
            ps.append([env[i:i + 30000]])
    w.save(dst)
    return rows


tmp = tempfile.mkdtemp(prefix="df-pricing-contract-")
try:
    wb_path = os.path.join(tmp, "fixture.xlsx")
    rows = _fixture_workbook(wb_path, fx_pricing, fx_fin)
    check("fixture envelope was chunked across multiple rows", rows > 3, f"rows={rows}")
    twb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    try:
        rebuilt = converter.build_pricing(twb)
        check("converter build_pricing reassembles the fixture exactly", rebuilt == fx_pricing)
        config = converter.build_store_config(twb)
    finally:
        twb.close()
    check("converter build_store_config carries the fixture as `pricing`",
          config.get("pricing") == fx_pricing)
    check("converter carries the fixture's authorized financing and untouched promotions",
          config.get("financing") == fx_fin and config.get("promotions") == cfg.get("promotions"))
    frep = validation.validate_financing(config, allowed_source_hosts=fin_hosts)
    check("the fixture's financing block is itself valid financing governance",
          frep.ok, "; ".join(frep.errors[:3]))
    rep = validation.validate_pricing(config, now=CLOCK, **KW)
    check("validate_pricing ADMITS the populated fixture under its clock (no errors, no warnings)",
          rep.ok and not rep.warnings, "; ".join(rep.errors[:3] + rep.warnings[:2]))

    # Non-vacuity: single-field corruptions of the SAME document, each refused,
    # one per reviewed contract rule.
    def corrupt(path, value):
        doc = json.loads(json.dumps(config))
        node = doc
        for k in path[:-1]:
            node = node[k]
        if value is KeyError:
            del node[path[-1]]
        else:
            node[path[-1]] = value
        return doc

    CASES = [
        ("displayEnabled flipped true", ("pricing", "displayEnabled"), True, "Phase 2.2"),
        ("product size removed (never inferred)", ("pricing", "products", 0, "size"), KeyError,
         "never inferred from another size"),
        ("amountMinor as a float (scope matched, so only the money rule can fire)",
         ("pricing", "products", 0, "price", "amountMinor"), 3699.0, "positive integer of minor units"),
        ("evidence status from the promotions ladder",
         ("pricing", "products", 0, "evidence", "status"), "retailer-full-page-archive",
         "evidence.status"),
        ("a transaction amount in configuration (§1)",
         ("pricing", "purchaseAssessment", "transactionAmountMinor"), 250000,
         "a transaction amount is a runtime argument, never configuration"),
        ("purchase policy = product price (§1)", ("pricing", "purchaseAssessment"),
         {"policy": "product-price"}, "a single product price is never it"),
        ("referenced plan unverified (§2)", ("financing", "plans", 0, "verified"), False,
         "referenced plan is not verified"),
        ("referenced plan stale (§2)", ("financing", "plans", 0, "verifiedAt"),
         "2026-07-31T16:43:00-05:00", "a stale plan admits no formula"),
        ("formula source outside the financing allowlist (§2)",
         ("pricing", "formulas", 0, "sourceUrl"), "https://linqcdn.avbportal.com/x",
         "canonical financing source-host allowlist"),
        ("price changed after clearance (§3)", ("pricing", "products", 0, "price", "amountMinor"),
         359900, "clearance.scope.amountMinor"),
        ("Queen -> King with the same SKU and price (§3 scope.size)",
         ("pricing", "products", 0, "size"), "king", "clearance.scope.size"),
        ("evidence classification changed (§3 scope.evidenceStatus)",
         ("pricing", "products", 0, "evidence", "status"), "retailer-price-list-current",
         "clearance.scope.evidenceStatus"),
        ("an end bound appears after clearance (§3 scope.windowEndsAt)",
         ("pricing", "products", 0, "window"), {"startAt": None, "endsAt": "2026-09-01T00:00:00-05:00"},
         "clearance.scope.windowEndsAt"),
        ("legal approval withdrawn (§4)", ("pricing", "presentation", "approvals", "legal"),
         {"status": "unapproved", "by": "", "at": None}, "one approval never stands in for another"),
        ("a disclosure loses its Spanish (§4)",
         ("pricing", "presentation", "disclosures", 0, "es"), "", "disclosures[0].es"),
        ("freshness policy unapproved while enabled (§5)", ("pricing", "freshness"),
         {"status": "unapproved", "maxAgeDays": None, "approvedBy": "", "approvedAt": None},
         "freshness.status must be 'approved'"),
        ("source policy unapproved while enabled (§5)", ("pricing", "sourcePolicy"),
         {"status": "unapproved", "allowedSourceHosts": [], "approvedBy": "", "approvedAt": None},
         "sourcePolicy.status must be 'approved'"),
    ]
    for label, path, value, needle in CASES:
        doc = corrupt(path, value)
        if path[-1] == "amountMinor" and isinstance(value, float):
            doc["pricing"]["products"][0]["clearance"]["scope"]["amountMinor"] = value
        bad = validation.validate_pricing(doc, now=CLOCK, **KW)
        check(f"fixture corrupted — {label} -> REFUSED",
              not bad.ok and any(needle in e for e in bad.errors),
              "; ".join(bad.errors[:2]) or "(admitted)")
    # mattress -> accessory with everything else re-pointed except the cleared kind
    kinddoc = json.loads(json.dumps(config))
    e0 = kinddoc["pricing"]["products"][0]
    e0.update({"productKind": "accessory", "productId": sorted(accessory_ids)[0], "size": None})
    e0["clearance"]["scope"].update({"productId": sorted(accessory_ids)[0], "size": None})
    kindrep = validation.validate_pricing(kinddoc, now=CLOCK, **KW)
    check("fixture corrupted — mattress -> accessory (§3 scope.productKind) -> REFUSED",
          not kindrep.ok and any("clearance.scope.productKind" in e for e in kindrep.errors),
          "; ".join(kindrep.errors[:2]) or "(admitted)")
    stale = validation.validate_pricing(config, now=CLOCK + timedelta(days=30), **KW)
    check("the injected clock is the clock: same fixture 30 days on -> stale, refused",
          not stale.ok and any("older than maxAgeDays" in e for e in stale.errors))
    fin_ok_price_bad = corrupt(("pricing", "formulas", 0, "sourceUrl"), "https://www.synchrony.com/terms")
    okrep = validation.validate_pricing(fin_ok_price_bad, now=CLOCK, **KW)
    check("formula source on a financing-only host is ADMITTED (financing policy, not the price list)",
          okrep.ok, "; ".join(okrep.errors[:2]))

    # ---- gate split (owner ruling 2026-08-28): the re-bound bindings ----------
    # The fixture already proves the dark formula validates with exact-term
    # output authorization OFF (production's own state). The flag is orthogonal:
    # switching it ON changes nothing for validation.
    on_doc = corrupt(("financing", "exactPromotionsEnabled"), True)
    onrep = validation.validate_pricing(on_doc, now=CLOCK, **KW)
    check("gate split — exact-term authorization ON is equally admitted (validation is orthogonal to output)",
          onrep.ok, "; ".join(onrep.errors[:2]))
    # Activation approvals absent while ENABLED and dark -> ADMITTED (the dark
    # resolver's case (a) configuration): MAP clearance not-cleared, blank
    # authority, presentation unapproved/pending and EMPTY.
    ua_doc = json.loads(json.dumps(config))
    ua_doc["pricing"]["authority"] = {"owner": "", "role": ""}
    ua_doc["pricing"]["products"][0]["clearance"] = {
        "status": "not-cleared", "attestedBy": "", "attestedAt": None, "scope": None}
    ua_doc["pricing"]["presentation"] = {
        "status": "unapproved",
        "approvals": {"business": {"status": "unapproved", "by": "", "at": None},
                      "legal": {"status": "unapproved", "by": "", "at": None},
                      "nativeReview": {"status": "pending", "by": "", "at": None}},
        "assumptions": [], "disclosures": [], "states": {},
    }
    ua_rep = validation.validate_pricing(ua_doc, now=CLOCK, **KW)
    check("gate split — ENABLED dark with every activation approval absent -> ADMITTED (case (a) config)",
          ua_rep.ok, "; ".join(ua_rep.errors[:3]))
    # The SAME document at activation -> every re-bound approval error fires
    # alongside the permanent display lock, so the approvals were MOVED to the
    # activation set, never dropped.
    ua_doc["pricing"]["displayEnabled"] = True
    act_rep = validation.validate_pricing(ua_doc, now=CLOCK, **KW)
    act_needles = ("Phase 2.2", "authority.owner",
                   "clearance.status must be 'cleared' or an attested",
                   "presentation.status must be 'approved' at activation",
                   "disclosures must be non-empty at activation")
    for needle in act_needles:
        check(f"gate split — same document at activation -> refused, names {needle[:44]!r}",
              any(needle in e for e in act_rep.errors), "; ".join(act_rep.errors[:3]))

    # The real converter subprocess refuses a hostile Pricing payload without a
    # traceback (its freshness verdict depends on the wall clock, so only the
    # shape is asserted through the subprocess).
    hostile = os.path.join(tmp, "hostile.xlsx")
    _fixture_workbook(hostile, {"products": "bad", "displayEnabled": True})
    proc = subprocess.run([sys.executable, os.path.join(REPO, "tools", "convert_store_data.py"),
                           hostile, "--validate-only"], capture_output=True, text=True, cwd=REPO)
    out = (proc.stdout or "") + (proc.stderr or "")
    check("converter REFUSES a hostile Pricing payload (nonzero exit)", proc.returncode != 0)
    check("converter names the pricing defects, no traceback",
          "pricing" in out and "displayEnabled" in out and "Traceback" not in out,
          out.strip()[-160:])
finally:
    shutil.rmtree(tmp, ignore_errors=True)

# ---- mutation-sweep find strings match exactly once ---------------------------
print("Mutation-sweep find-string uniqueness (the sweep does not check this):")
SWEEP_FINDS = [
    ("tools/validation.py",
     "    if display is True:\n        r.add_error(\"pricing.displayEnabled must be false"),
    ("tools/validation.py",
     "                elif type(size) is not str or size not in size_set:"),
    ("tools/validation.py",
     "        if enabled:\n            r.add_error(msg)\n        else:\n            r.add_warning(msg)"),
    ("tools/validation.py",
     "                if type(amt) is not int or amt <= 0 or amt > PRICING_AMOUNT_MINOR_MAX:\n"
     "                    r.add_error(f\"{tag}.price.amountMinor"),
    ("tools/validation.py",
     "    for k in obj:\n        if k not in allowed:\n            r.add_error(f\"{tag}: key"),
    ("data/store-config.json",
     "    \"enabled\": false,\n    \"displayEnabled\": false,"),
    ("tools/validation.py",
     "    \"transactionamountminor\", \"transactionamount\", \"purchaseamountminor\","),
    ("tools/validation.py",
     "        if not fin_enabled:\n            r.add_error(\"pricing.formulas: financing must be enabled — a formula \""),
    ("tools/validation.py",
     "                            elif scope[key] != expected[key] or type(scope[key]) is not type(expected[key]):"),
    ("tools/validation.py",
     "        if activation and not papproved:\n            r.add_error(\"pricing.presentation.status must be 'approved' at activation \""),
    ("tools/validation.py",
     "        if enabled and not approved:\n            r.add_error(\"pricing.freshness.status must be 'approved' (with maxAgeDays, \""),
    ("tools/validation.py",
     "                            \"size\": size if kind == \"mattress\" else None,"),
    # ---- Phase 2.1b resolver anchors (index.html) and the payload anchor ----
    ("index.html",
     "        eligible = p.displayEnabled === true"),
    ("index.html",
     "          freshness = (now - evInstant) > mad * 86400000 ? 'stale' : 'fresh';"),
    ("index.html",
     "      var txn = q.transactionAmountMinor;"),
    ("index.html",
     "          threshold = txn >= minMajor * 100 ? 'met' : 'not-met';"),
    ("index.html",
     "        calculation = 'quote-only';"),
    ("index.html",
     "      const payload = {\n        storeName: (STORE_CONFIG && STORE_CONFIG.storeName) || '',"),
    ("index.html",
     "        var price = Number(primary.price) > 0\n"
     "          ? sleepSystemText({ en: 'From $', es: 'Desde $' }) + Number(primary.price).toLocaleString()"),
]
for target, find in SWEEP_FINDS:
    n = _lf(_read(target)).count(find)
    check(f"{target}: sweep find string matches exactly once ({find[:48]!r}...)", n == 1, f"count={n}")

print(f"\nPricing contract check: {passed} passed, {failed} failed")
sys.exit(1 if failed else 0)
